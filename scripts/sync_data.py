#!/usr/bin/env python3
import os
import re
import sys
import json
from pathlib import Path
import openpyxl

def get_bilingual_name(name_val):
    if name_val is None:
        return ""
    name = str(name_val).strip()
    name_map = {
        "Minisaw": "미니톱 (Minisaw)",
        "저울": "저울 (Scale)",
        "미세저울": "미세저울 (Microscale)",
        "가열교반기": "가열교반기 (Hot Plate Stirrer)",
        "pH/이온 측정기/ 비색계": "pH/이온 측정기/비색계 (pH/Ion Meter & Colorimeter)",
        "접촉각": "접촉각 측정기 (Contact Angle Meter)",
        "Spin Coater": "스핀 코터 (Spin Coater)",
        "UTM": "만능 재료 시험기 (UTM)",
        "incubator": "인큐베이터 (Incubator)",
        "형광 단백질 광원": "형광 단백질 광원 (Fluorescence Light Source)",
        "Mimics S/W": "미믹스 소프트웨어 (Mimics S/W)",
        "광조사기": "광조사기 (Light Curing Unit)",
        "멸균기": "고압멸균기 (Autoclave)",
        "Vacuum oven": "진공 오븐 (Vacuum Oven)",
        "Epoch": "에포크 마이크로플레이트 리더 (Epoch Microplate Reader)",
        "초음파 세척기": "초음파 세척기 (Ultrasonic Cleaner)",
        "Thermo CO2 incubator": "써모 CO2 인큐베이터 (Thermo CO2 Incubator)",
        "질소탱크": "질소탱크 (Nitrogen Tank)",
        "Real-Time PCR": "실시간 유전자 증폭기 (Real-Time PCR)",
        "Centrifuge": "원심분리기 (Centrifuge)",
        "thermal cycler": "열 순환기 (Thermal Cycler)",
        "phcbi CO2 incubator": "phcbi CO2 인큐베이터 (phcbi CO2 Incubator)",
        "uniz": "유니즈 3D 프린터 (UNIZ 3D Printer)",
        "질소 경화기": "질소 경화기 (Nitrogen Curing Chamber)",
        "Spining": "일렉트로스피닝 장치 (Spinning Device)",
        "회전증발 농축기": "회전증발 농축기 (Rotary Evaporator)",
        "ultrasonic homogenizer probes": "초음파 균질기 프로브 (Ultrasonic Homogenizer Probes)",
        "UV 경화기 2개": "UV 경화기 - 2개 (UV Curing Units - 2pcs)",
        "큐레이": "큐레이 구강카메라 (Q-ray Oral Camera)",
        "electrospinning": "전기방사 장치 (Electrospinning)",
        "구강 스캐너": "구강 스캐너 (Intraoral Scanner)",
        "FT-IR": "적외선 분광 광도계 (FT-IR)",
        "연마기": "연마기 (Polishing Machine)",
        "자동 제습 보관함": "자동 제습 보관함 (Auto-Desiccating Cabinet)",
        "진공 데시게이터": "진공 데시게이터 (Vacuum Desiccator)",
        "EVOS M5000 이미징시스템": "EVOS M5000 이미징 시스템 (EVOS M5000 Imaging System)",
        "볼텍스 믹서": "볼텍스 믹서 (Vortex Mixer)",
        "냉장고 (세균, 공용)": "공용 세균 냉장고 (Shared Refrigerator)",
        "사진촬영": "암실 사진 촬영기 (Photo Documentation System)",
        "전동 파이펫, 파이펫": "전동 파이펫 및 피펫 세트 (Pipette Aid & Pipette Set)",
        "현미경": "광학 현미경 (Microscope)",
        "교정용 Vac Plus": "교정용 백플러스 (Calibration Vac Plus)",
        "PC 본체": "데스크탑 PC 본체 (PC Desktop Tower)",
        "폐쇄기": "씰링 튜브 폐쇄기 (Tube Sealer)",
        "프린터": "사무용 프린터 (Office Printer)",
        "노트북": "노트북 컴퓨터 (Laptop Computer)",
        "마이크로칩": "마이크로플루이딕 마이크로칩 (Microfluidic Microchip)",
        "펌프 장비": "유체 주입 펌프 장비 (Syringe Pump Equipment)",
        "UV 멸균": "UV 멸균기 (UV Sterilizer)"
    }
    return name_map.get(name, name)

def sync():
    root = Path(__file__).resolve().parents[1]
    excel_path = root / "담당자.xlsx"
    html_path = root / "index.html"
    
    if not excel_path.exists():
        print(f"Error: Excel file '{excel_path}' not found.")
        return False
        
    print(f"Reading general equipment from {excel_path.name}...")
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheet_name = "장비 목록 (전체)"
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in Excel file.")
            print(f"Available sheets: {wb.sheetnames}")
            return False
            
        ws = wb[sheet_name]
        inventory_data = []
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            if len(row) < 7:
                continue
                
            no_val = row[1]
            name_val = row[2]
            loc_val = row[4]
            mgr_val = row[5]
            remarks_val = row[6]
            
            if no_val is None:
                continue
                
            try:
                no = int(no_val)
            except (ValueError, TypeError):
                continue
                
            if not name_val:
                continue
                
            loc = str(loc_val).strip() if loc_val is not None else ""
            if loc == "미지정":
                loc = ""
                
            mgr = str(mgr_val).strip() if mgr_val is not None else ""
            if mgr == "미지정":
                mgr = ""
                
            remarks = str(remarks_val).strip() if remarks_val is not None else ""
            
            inventory_data.append({
                "no": no,
                "name": get_bilingual_name(name_val),
                "loc": loc,
                "mgr": mgr,
                "remarks": remarks
            })
            
        wb.close()
        print(f"Successfully parsed {len(inventory_data)} general equipment entries.")
        
    except Exception as e:
        print(f"An error occurred while reading 담당자.xlsx: {e}")
        return False

    # Now read pipette and PC from archive
    pipette_data = []
    pc_data = []
    
    archive_dir = root / "archive"
    archive_files = []
    if archive_dir.exists():
        archive_files = sorted(
            [f for f in archive_dir.glob("*_장비 보유 현황.xlsx") if f.is_file()],
            key=lambda x: x.name
        )
        
    if not archive_files:
        print("Warning: No matching *_장비 보유 현황.xlsx found in archive/ folder. Skipping Pipette and PC sync.")
    else:
        latest_archive = archive_files[-1]
        print(f"Reading Pipette & PC from latest archive: {latest_archive.name}...")
        try:
            wb_arc = openpyxl.load_workbook(latest_archive, read_only=True, data_only=True)
            
            # Parse Pipette
            if "Pipette" in wb_arc.sheetnames:
                ws_pip = wb_arc["Pipette"]
                # Headers are on row 2, data starts on row 3
                for row in ws_pip.iter_rows(min_row=3, values_only=True):
                    if not row or len(row) < 5:
                        continue
                    no_val = row[0]
                    name_val = row[1]
                    spec_val = row[2]
                    loc_val = row[3]
                    brand_val = row[4]
                    calib_val = row[5] if len(row) > 5 else None
                    broken_val = row[6] if len(row) > 6 else None
                    remarks_val = row[7] if len(row) > 7 else None
                    
                    if no_val is None:
                        continue
                    try:
                        no = int(no_val)
                    except (ValueError, TypeError):
                        continue
                    if not name_val:
                        continue
                        
                    loc = str(loc_val).strip() if loc_val is not None else ""
                    if loc == "미지정":
                        loc = ""
                    brand = str(brand_val).strip() if brand_val is not None else ""
                    calib = str(calib_val).strip() if calib_val is not None else ""
                    broken = str(broken_val).strip() if broken_val is not None else ""
                    remarks = str(remarks_val).strip() if remarks_val is not None else ""
                    
                    pipette_data.append({
                        "no": no,
                        "name": str(name_val).strip(),
                        "spec": spec_val or "",
                        "loc": loc,
                        "brand": brand,
                        "calib": calib,
                        "broken": broken,
                        "remarks": remarks
                    })
                print(f"Successfully parsed {len(pipette_data)} pipette entries.")
            else:
                print("Warning: 'Pipette' sheet not found in the archive Excel.")
                
            # Parse PC
            if "PC" in wb_arc.sheetnames:
                ws_pc = wb_arc["PC"]
                # Headers are on row 2, data starts on row 3
                for row in ws_pc.iter_rows(min_row=3, values_only=True):
                    if not row or len(row) < 5:
                        continue
                    no_val = row[0]
                    kind_val = row[1]
                    label_val = row[2]
                    loc_val = row[3]
                    user_val = row[4]
                    brand_val = row[5]
                    photo_val = row[6] if len(row) > 6 else None
                    sticker_val = row[7] if len(row) > 7 else None
                    
                    if no_val is None:
                        continue
                    try:
                        no = int(no_val)
                    except (ValueError, TypeError):
                        continue
                    if not kind_val:
                        continue
                        
                    loc = str(loc_val).strip() if loc_val is not None else ""
                    if loc == "미지정":
                        loc = ""
                    kind = str(kind_val).strip()
                    label = str(label_val).strip() if label_val is not None else ""
                    user = str(user_val).strip() if user_val is not None else ""
                    brand = str(brand_val).strip() if brand_val is not None else ""
                    photo = str(photo_val).strip() if photo_val is not None else ""
                    sticker = str(sticker_val).strip() if sticker_val is not None else ""
                    
                    pc_data.append({
                        "no": no,
                        "kind": kind,
                        "label": label,
                        "loc": loc,
                        "user": user,
                        "brand": brand,
                        "photo": photo,
                        "sticker": sticker
                    })
                print(f"Successfully parsed {len(pc_data)} PC entries.")
            else:
                print("Warning: 'PC' sheet not found in the archive Excel.")
                
            wb_arc.close()
        except Exception as e:
            print(f"An error occurred while reading the archive Excel: {e}")
            # Keep empty lists to avoid breaking the script
            
    # Write to index.html
    if not html_path.exists():
        print(f"Error: index.html not found at {html_path}")
        return False
        
    try:
        html_content = html_path.read_text(encoding="utf-8")
        
        # Helper to format data
        def format_json(data_list):
            formatted_json = "[\n"
            for i, item in enumerate(data_list):
                comma = "," if i < len(data_list) - 1 else ""
                item_json = json.dumps(item, ensure_ascii=False)
                formatted_json += f"      {item_json}{comma}\n"
            formatted_json += "    ]"
            return formatted_json
            
        # Replace inventoryData
        formatted_inventory = format_json(inventory_data)
        pattern_inv = r"(const inventoryData = )\[[\s\S]*?\];"
        replacement_inv = f"\\1{formatted_inventory};"
        html_content, count_inv = re.subn(pattern_inv, replacement_inv, html_content)
        if count_inv == 0:
            print("Error: Could not locate 'const inventoryData = [...];' in index.html.")
            return False
            
        # Replace pipetteData
        formatted_pipette = format_json(pipette_data)
        pattern_pip = r"(const pipetteData = )\[[\s\S]*?\];"
        replacement_pip = f"\\1{formatted_pipette};"
        html_content, count_pip = re.subn(pattern_pip, replacement_pip, html_content)
        if count_pip == 0:
            print("Warning: Could not locate 'const pipetteData = [...];' in index.html (will be added).")
            
        # Replace pcData
        formatted_pc = format_json(pc_data)
        pattern_pc = r"(const pcData = )\[[\s\S]*?\];"
        replacement_pc = f"\\1{formatted_pc};"
        html_content, count_pc = re.subn(pattern_pc, replacement_pc, html_content)
        if count_pc == 0:
            print("Warning: Could not locate 'const pcData = [...];' in index.html (will be added).")
            
        html_path.write_text(html_content, encoding="utf-8")
        print("index.html has been successfully updated with new data.")
        return True
        
    except Exception as e:
        print(f"An error occurred while updating index.html: {e}")
        return False

if __name__ == "__main__":
    success = sync()
    if not success:
        sys.exit(1)
