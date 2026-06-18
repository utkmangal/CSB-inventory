import openpyxl
from pathlib import Path
from io import BytesIO
from PIL import Image

def main():
    root = Path(__file__).resolve().parents[1]
    xlsx_path = root / "archive" / "20260625_장비 보유 현황.xlsx"
    images_dir = root / "images"
    images_dir.mkdir(exist_ok=True)
    
    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found.")
        return
        
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['PC']
    
    print(f"Loaded {xlsx_path.name}. Extracting images from 'PC' sheet...")
    
    extracted_photos = 0
    extracted_stickers = 0
    
    for idx, img in enumerate(ws._images):
        anchor = img.anchor
        if not hasattr(anchor, '_from'):
            continue
            
        row_idx = anchor._from.row + 1
        col_idx = anchor._from.col + 1
        
        no_val = ws.cell(row=row_idx, column=1).value
        if no_val is None:
            no_val = ws.cell(row=row_idx-1, column=1).value
            if no_val is None:
                no_val = ws.cell(row=row_idx+1, column=1).value
                
        if no_val is None:
            print(f"Warning: Could not find 'No.' for image {idx} at row {row_idx}")
            continue
            
        try:
            no_val = int(no_val)
        except (ValueError, TypeError):
            continue
            
        try:
            if isinstance(img.ref, BytesIO):
                img_bytes = img.ref.getvalue()
            elif hasattr(img.ref, 'read'):
                img.ref.seek(0)
                img_bytes = img.ref.read()
            else:
                print(f"Unknown image ref type: {type(img.ref)}")
                continue
                
            pil_img = Image.open(BytesIO(img_bytes))
            if pil_img.mode in ('RGBA', 'P', 'LA'):
                pil_img = pil_img.convert('RGB')
            elif pil_img.mode != 'RGB':
                pil_img = pil_img.convert('RGB')
                
            if col_idx in (6, 7):
                dest_path = images_dir / f"pc_{no_val}.jpeg"
                pil_img.save(dest_path, "JPEG", quality=90)
                print(f"Saved photo: pc_{no_val}.jpeg (anchored at row {row_idx}, col {col_idx})")
                extracted_photos += 1
            elif col_idx == 8:
                dest_path = images_dir / f"pc_{no_val}_sticker.jpeg"
                pil_img.save(dest_path, "JPEG", quality=90)
                print(f"Saved sticker: pc_{no_val}_sticker.jpeg (anchored at row {row_idx}, col {col_idx})")
                extracted_stickers += 1
            else:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"Image {idx} at row {row_idx}, col {col_letter} ignored (not F, G, or H)")
                
        except Exception as e:
            print(f"Error extracting image {idx} at row {row_idx}: {e}")
            
    print(f"\nExtraction complete: Saved {extracted_photos} photos and {extracted_stickers} stickers.")
    wb.close()

if __name__ == "__main__":
    main()
